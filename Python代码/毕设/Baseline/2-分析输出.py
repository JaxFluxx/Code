# -*- coding: utf-8 -*-
"""
From cleaned CODEX sample tables -> pair_table + cohort analysis + stratified analysis
Mac version

依赖第一段脚本输出：
/Users/jia/Desktop/学习 /毕业设计/实验/CODEX_clean_baseline/

输出：
/Users/jia/Desktop/学习 /毕业设计/实验/CODEX_pair_results_r20_r30_stratified/
  - pair_table.tsv
  - cohort_analysis.tsv
  - pair_table_with_group.tsv
  - pair_table_with_group.xlsx
  - stratified_stats.tsv
  - stratified_stats.xlsx
  - stratified_correlations.tsv
  - stratified_correlations.xlsx
  - plots_pairs/
  - plots/

说明：
- diff_* = Met - Primary
- stratified 部分已包含你要求的排序：
    R-only -> L-only -> QT-only -> 其他
"""

from pathlib import Path
import numpy as np
import pandas as pd
from scipy.stats import wilcoxon, spearmanr
from scipy.spatial import cKDTree
import matplotlib.pyplot as plt
from matplotlib import rcParams
from matplotlib.lines import Line2D
import shutil
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# =========================
# 0) PATHS
# =========================
EXPERIMENT_ROOT = Path("/Users/jia/Desktop/学习 /毕业设计/实验")
CLEAN_ROOT = EXPERIMENT_ROOT / "CODEX_clean_baseline"
PAIR_MANIFEST_PATH = CLEAN_ROOT / "pair_manifest.tsv"

OUT_DIR = EXPERIMENT_ROOT / "CODEX_pair_results_r20_r30_stratified"
PAIR_PLOT_DIR = OUT_DIR / "plots_pairs"
STRAT_PLOT_DIR = OUT_DIR / "plots"

# overwrite output dir
if OUT_DIR.exists():
    shutil.rmtree(OUT_DIR)
PAIR_PLOT_DIR.mkdir(parents=True, exist_ok=True)
STRAT_PLOT_DIR.mkdir(parents=True, exist_ok=True)

if not PAIR_MANIFEST_PATH.exists():
    raise FileNotFoundError(f"找不到 pair_manifest.tsv：{PAIR_MANIFEST_PATH}")

# matplotlib 中文字体回退
rcParams["font.sans-serif"] = [
    "PingFang SC",
    "Hiragino Sans GB",
    "Microsoft YaHei",
    "SimHei",
    "Noto Sans CJK SC",
    "Arial Unicode MS",
    "DejaVu Sans",
]
rcParams["axes.unicode_minus"] = False

# =========================
# 1) EXCEL FORMAT HELPERS
# =========================
YELLOW_FILL = PatternFill(fill_type="solid", fgColor="FFF59D")
HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")

def is_percent_col(col_name: str) -> bool:
    name = str(col_name).lower()
    return any(k in name for k in ["frac", "ratio", "within_"])

def excel_number_format_for_col(col_name: str) -> str:
    if is_percent_col(col_name):
        return "0.00%"
    return "0.00"

def safe_round_for_export(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    for c in out.columns:
        if pd.api.types.is_numeric_dtype(out[c]):
            def _fmt_num(x):
                if pd.isna(x):
                    return np.nan
                x = float(x)
                ax = abs(x)
                if ax == 0:
                    return 0.0
                if ax < 0.01 or ax >= 10000:
                    return float(f"{x:.2g}")   # 两位有效数字
                return round(x, 2)
            out[c] = out[c].map(_fmt_num)
    return out

def autofit_worksheet(ws, df: pd.DataFrame):
    for idx, col in enumerate(df.columns, start=1):
        max_len = len(str(col))
        for row in ws.iter_rows(min_row=2, min_col=idx, max_col=idx):
            val = row[0].value
            if val is None:
                continue
            max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(idx)].width = min(max_len + 2, 26)

def format_excel_file(xlsx_path: Path, df: pd.DataFrame, diff_yellow=True):
    wb = load_workbook(xlsx_path)
    ws = wb.active

    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    col_names = list(df.columns)
    for j, col_name in enumerate(col_names, start=1):
        num_fmt = excel_number_format_for_col(col_name)
        is_numeric = pd.api.types.is_numeric_dtype(df[col_name])

        for i in range(2, ws.max_row + 1):
            cell = ws.cell(row=i, column=j)
            if is_numeric and isinstance(cell.value, (int, float)):
                cell.number_format = num_fmt

            if diff_yellow and str(col_name).startswith("diff_"):
                cell.fill = YELLOW_FILL

        if diff_yellow and str(col_name).startswith("diff_"):
            ws.cell(row=1, column=j).fill = YELLOW_FILL
            ws.cell(row=1, column=j).font = Font(bold=True)

    ws.freeze_panes = "A2"
    autofit_worksheet(ws, df)
    wb.save(xlsx_path)

def save_tsv_and_xlsx(df: pd.DataFrame, tsv_path: Path, xlsx_path: Path):
    export_df = safe_round_for_export(df)
    export_df.to_csv(tsv_path, sep="\t", index=False)
    export_df.to_excel(xlsx_path, index=False)
    format_excel_file(xlsx_path, export_df, diff_yellow=True)

# =========================
# 2) HELPERS
# =========================
def met_group(met_sample: str) -> str:
    s = str(met_sample)
    if s.endswith("L"):
        return "L-only"
    if s.endswith("R"):
        return "R-only"
    if s.endswith("QT"):
        return "QT-only"
    return "Unknown"

def ovary_group(met_sample: str) -> str:
    s = str(met_sample)
    if s.endswith("L") or s.endswith("R"):
        return "Ovary(L/R)"
    if s.endswith("QT"):
        return "QT-only"
    return "Unknown"

def extract_sample_num(sample_name: str) -> float:
    s = str(sample_name)
    num = ""
    for ch in s:
        if ch.isdigit():
            num += ch
        else:
            break
    return float(num) if num else np.inf

def safe_div(a, b):
    if b is None or b == 0 or pd.isna(b):
        return np.nan
    return a / b

def wilcoxon_safe(x: np.ndarray):
    x = x[~np.isnan(x)]
    if len(x) < 2:
        return np.nan
    if np.allclose(x, 0):
        return 1.0
    try:
        return float(wilcoxon(x, zero_method="pratt").pvalue)
    except Exception:
        try:
            return float(wilcoxon(x).pvalue)
        except Exception:
            return np.nan

def spearman_safe(a, b):
    mask = (~np.isnan(a)) & (~np.isnan(b))
    if mask.sum() < 3:
        return np.nan, np.nan, int(mask.sum())
    rho, p = spearmanr(a[mask], b[mask])
    return float(rho), float(p), int(mask.sum())

def fmt(x):
    if pd.isna(x):
        return "NA"
    try:
        return f"{float(x):.4g}"
    except Exception:
        return str(x)

METRIC_CN = {
    "enrich_30_Tumor_M1": "Tumor-M1富集(30)",
    "enrich_20_Tumor_M1": "Tumor-M1富集(20)",
    "within_30_M2": "M2邻域占比(30)",
    "within_20_M2": "M2邻域占比(20)",
    "m2_frac_of_mac": "M2/巨噬比例",
    "tumor_neighbor_M1_ratio": "Tumor近邻M1比例",
    "tumor_neighbor_M2_ratio": "Tumor近邻M2比例",
}

def metric_cn(metric: str) -> str:
    return METRIC_CN.get(metric, metric)

def get_coords(df: pd.DataFrame, cell_type: str):
    sub = df[df["cell_type"] == cell_type]
    if len(sub) == 0:
        return np.empty((0, 2), dtype=float)
    return sub[["Center X", "Center Y"]].to_numpy(dtype=float)

def nearest_distances(src_coords: np.ndarray, tgt_coords: np.ndarray):
    if len(src_coords) == 0 or len(tgt_coords) == 0:
        return np.array([], dtype=float)
    tree = cKDTree(tgt_coords)
    d, _ = tree.query(src_coords, k=1)
    d = np.asarray(d, dtype=float)
    return d[np.isfinite(d)]

def dist_summary(arr: np.ndarray):
    if len(arr) == 0:
        return {"median": np.nan, "p25": np.nan, "p75": np.nan}
    return {
        "median": float(np.median(arr)),
        "p25": float(np.quantile(arr, 0.25)),
        "p75": float(np.quantile(arr, 0.75)),
    }

def within_r_ratio(src_coords: np.ndarray, tgt_coords: np.ndarray, r: float):
    if len(src_coords) == 0 or len(tgt_coords) == 0:
        return np.nan
    tree = cKDTree(tgt_coords)
    d, _ = tree.query(src_coords, k=1)
    d = np.asarray(d, dtype=float)
    return float(np.mean(d <= r))

def tumor_neighbor_ratios(tumor_coords: np.ndarray, m1_coords: np.ndarray, m2_coords: np.ndarray, k=10):
    if len(tumor_coords) == 0:
        return np.nan, np.nan
    mac_coords = []
    labels = []

    if len(m1_coords) > 0:
        mac_coords.append(m1_coords)
        labels.extend(["M1"] * len(m1_coords))
    if len(m2_coords) > 0:
        mac_coords.append(m2_coords)
        labels.extend(["M2"] * len(m2_coords))

    if len(mac_coords) == 0:
        return np.nan, np.nan

    mac_coords = np.vstack(mac_coords)
    labels = np.array(labels)
    tree = cKDTree(mac_coords)

    k_eff = min(k, len(mac_coords))
    _, idx = tree.query(tumor_coords, k=k_eff)

    if k_eff == 1:
        idx = idx.reshape(-1, 1)

    m1_ratios = []
    m2_ratios = []

    for row in idx:
        lab = labels[row]
        m1_ratios.append(np.mean(lab == "M1"))
        m2_ratios.append(np.mean(lab == "M2"))

    return float(np.mean(m1_ratios)), float(np.mean(m2_ratios))

def avg_count_within_radius(center_coords: np.ndarray, target_coords: np.ndarray, r: float):
    """
    对每个 center 点，在半径 r 内数 target 邻居个数，再对所有 center 取平均
    """
    if len(center_coords) == 0 or len(target_coords) == 0:
        return np.nan
    tree = cKDTree(target_coords)
    counts = tree.query_ball_point(center_coords, r=r, return_length=True)
    counts = np.asarray(counts, dtype=float)
    return float(np.mean(counts))

def avg_m1_m2_interaction(m1_coords: np.ndarray, m2_coords: np.ndarray, r: float):
    """
    以 M1 为中心，看半径 r 内平均能数到多少 M2
    """
    if len(m1_coords) == 0 or len(m2_coords) == 0:
        return np.nan
    tree = cKDTree(m2_coords)
    counts = tree.query_ball_point(m1_coords, r=r, return_length=True)
    counts = np.asarray(counts, dtype=float)
    return float(np.mean(counts))

def compute_sample_metrics(df: pd.DataFrame):
    tumor_coords = get_coords(df, "Tumor")
    m1_coords = get_coords(df, "M1")
    m2_coords = get_coords(df, "M2")

    tumor_n = len(tumor_coords)
    m1_n = len(m1_coords)
    m2_n = len(m2_coords)

    m2_frac = safe_div(m2_n, (m1_n + m2_n))

    m1_d = nearest_distances(m1_coords, tumor_coords)
    m2_d = nearest_distances(m2_coords, tumor_coords)

    m1_dist = dist_summary(m1_d)
    m2_dist = dist_summary(m2_d)

    metrics = {
        "tumor_n": tumor_n,
        "m1_n": m1_n,
        "m2_n": m2_n,
        "m2_frac_of_mac": m2_frac,

        "m1_dist_median": m1_dist["median"],
        "m2_dist_median": m2_dist["median"],
        "delta_median": (
            m1_dist["median"] - m2_dist["median"]
            if pd.notna(m1_dist["median"]) and pd.notna(m2_dist["median"])
            else np.nan
        ),

        "within_20_M1": within_r_ratio(m1_coords, tumor_coords, 20),
        "within_20_M2": within_r_ratio(m2_coords, tumor_coords, 20),
        "within_30_M1": within_r_ratio(m1_coords, tumor_coords, 30),
        "within_30_M2": within_r_ratio(m2_coords, tumor_coords, 30),

        "tumor_neighbor_M1_ratio": np.nan,
        "tumor_neighbor_M2_ratio": np.nan,

        "enrich_20_Tumor_M1": avg_count_within_radius(tumor_coords, m1_coords, 20),
        "enrich_20_Tumor_M2": avg_count_within_radius(tumor_coords, m2_coords, 20),
        "enrich_20_M1_M2": avg_m1_m2_interaction(m1_coords, m2_coords, 20),

        "enrich_30_Tumor_M1": avg_count_within_radius(tumor_coords, m1_coords, 30),
        "enrich_30_Tumor_M2": avg_count_within_radius(tumor_coords, m2_coords, 30),
        "enrich_30_M1_M2": avg_m1_m2_interaction(m1_coords, m2_coords, 30),
    }

    neigh_m1, neigh_m2 = tumor_neighbor_ratios(tumor_coords, m1_coords, m2_coords, k=10)
    metrics["tumor_neighbor_M1_ratio"] = neigh_m1
    metrics["tumor_neighbor_M2_ratio"] = neigh_m2

    return metrics, {"m1_dists": m1_d, "m2_dists": m2_d}

def paired_line_plot(sub_df: pd.DataFrame, metric: str, out_png: Path, title: str):
    pcol = f"primary_{metric}"
    mcol = f"met_{metric}"
    if pcol not in sub_df.columns or mcol not in sub_df.columns:
        return
    tmp = sub_df[["patient_id", pcol, mcol]].dropna()
    if len(tmp) == 0:
        return

    plt.figure(figsize=(5.2, 4.0))
    pvals = []
    mvals = []
    for _, r in tmp.iterrows():
        pv = float(r[pcol])
        mv = float(r[mcol])
        pvals.append(pv)
        mvals.append(mv)
        color = "#d62728" if mv >= pv else "#1f77b4"  # 红: 上升, 蓝: 下降
        plt.plot([0, 1], [pv, mv], marker="o", linewidth=1.6, color=color, alpha=0.85)

    # 叠加组均值变化线，便于整体方向判断
    mean_p = float(np.mean(pvals))
    mean_m = float(np.mean(mvals))
    plt.plot([0, 1], [mean_p, mean_m], marker="D", markersize=5.5, linewidth=2.4, color="#111111")
    plt.xticks([0, 1], ["原发灶", "转移灶"])
    plt.title(title)
    plt.ylabel(metric_cn(metric))
    legend_handles = [
        Line2D([0], [0], color="#1f77b4", marker="o", lw=1.8, label="单对样本：转移<原发"),
        Line2D([0], [0], color="#d62728", marker="o", lw=1.8, label="单对样本：转移>=原发"),
        Line2D([0], [0], color="#111111", marker="D", lw=2.4, label="组均值变化"),
    ]
    plt.legend(handles=legend_handles, frameon=False, loc="best", fontsize=8.8)
    plt.tight_layout()
    plt.savefig(out_png, dpi=200)
    plt.close()

def diff_hist(sub_df: pd.DataFrame, metric: str, out_png: Path, title: str):
    dcol = f"diff_{metric}"
    if dcol not in sub_df.columns:
        return
    x = sub_df[dcol].dropna().values
    if len(x) == 0:
        return

    plt.figure(figsize=(5.2, 4.0))
    plt.hist(x, bins=min(20, max(5, len(x))), edgecolor="black")
    plt.axvline(0, linestyle="--")
    plt.title(title)
    plt.xlabel(f"diff_{metric} (Met - Primary)")
    plt.ylabel("Count")
    plt.tight_layout()
    plt.savefig(out_png, dpi=200)
    plt.close()

def pair_density_plot(primary_raw, met_raw, out_png: Path, title: str):
    plt.figure(figsize=(6.8, 4.8))
    plotted = False

    for arr, label in [
        (primary_raw.get("m1_dists", np.array([])), "Primary-M1"),
        (primary_raw.get("m2_dists", np.array([])), "Primary-M2"),
        (met_raw.get("m1_dists", np.array([])), "Met-M1"),
        (met_raw.get("m2_dists", np.array([])), "Met-M2"),
    ]:
        arr = np.asarray(arr, dtype=float)
        arr = arr[np.isfinite(arr)]
        arr = arr[arr > 0]
        if len(arr) > 0:
            plt.hist(arr, bins=30, density=True, histtype="step", linewidth=1.5, label=label)
            plotted = True

    if plotted:
        plt.legend()
        plt.title(title)
        plt.xlabel("Nearest distance to Tumor")
        plt.ylabel("Density")
        plt.tight_layout()
        plt.savefig(out_png, dpi=200)
    plt.close()

def pair_boxplot(primary_raw, met_raw, out_png: Path, title: str):
    data = []
    labels = []

    for arr, label in [
        (primary_raw.get("m1_dists", np.array([])), "P-M1"),
        (primary_raw.get("m2_dists", np.array([])), "P-M2"),
        (met_raw.get("m1_dists", np.array([])), "M-M1"),
        (met_raw.get("m2_dists", np.array([])), "M-M2"),
    ]:
        arr = np.asarray(arr, dtype=float)
        arr = arr[np.isfinite(arr)]
        if len(arr) > 0:
            data.append(arr)
            labels.append(label)

    if len(data) == 0:
        return

    plt.figure(figsize=(6.4, 4.4))
    plt.boxplot(data, labels=labels, showfliers=False)
    plt.title(title)
    plt.ylabel("Nearest distance to Tumor")
    plt.tight_layout()
    plt.savefig(out_png, dpi=200)
    plt.close()

def pair_within_r_plot(primary_metrics, met_metrics, out_png: Path, title: str):
    labels = ["M1@20", "M2@20", "M1@30", "M2@30"]
    pvals = [
        primary_metrics.get("within_20_M1", np.nan),
        primary_metrics.get("within_20_M2", np.nan),
        primary_metrics.get("within_30_M1", np.nan),
        primary_metrics.get("within_30_M2", np.nan),
    ]
    mvals = [
        met_metrics.get("within_20_M1", np.nan),
        met_metrics.get("within_20_M2", np.nan),
        met_metrics.get("within_30_M1", np.nan),
        met_metrics.get("within_30_M2", np.nan),
    ]

    x = np.arange(len(labels))
    w = 0.36

    plt.figure(figsize=(6.6, 4.4))
    plt.bar(x - w / 2, pvals, width=w, label="Primary")
    plt.bar(x + w / 2, mvals, width=w, label="Met")
    plt.xticks(x, labels)
    plt.ylim(bottom=0)
    plt.title(title)
    plt.ylabel("Ratio")
    plt.legend()
    plt.tight_layout()
    plt.savefig(out_png, dpi=200)
    plt.close()

# =========================
# 3) LOAD PAIRS
# =========================
pairs = pd.read_csv(PAIR_MANIFEST_PATH, sep="\t")
required_cols = ["patient_id", "primary_sample", "met_sample", "primary_analysis_path", "met_analysis_path"]
for col in required_cols:
    if col not in pairs.columns:
        raise ValueError(f"pair_manifest 缺少列：{col}")

print(f"共读取样本对：{len(pairs)}")

# =========================
# 4) PAIR-LEVEL ANALYSIS
# =========================
pair_rows = []

for _, row in pairs.iterrows():
    patient_id = row["patient_id"]
    primary_sample = row["primary_sample"]
    met_sample = row["met_sample"]

    primary_path = Path(row["primary_analysis_path"])
    met_path = Path(row["met_analysis_path"])

    if not primary_path.exists() or not met_path.exists():
        print(f"[跳过] 缺少clean表：{primary_sample} / {met_sample}")
        continue

    primary_df = pd.read_csv(primary_path, sep="\t")
    met_df = pd.read_csv(met_path, sep="\t")

    primary_metrics, primary_raw = compute_sample_metrics(primary_df)
    met_metrics, met_raw = compute_sample_metrics(met_df)

    one = {
        "patient_id": patient_id,
        "primary_sample": primary_sample,
        "met_sample": met_sample,
    }

    # primary / met / diff
    metric_names = sorted(set(primary_metrics.keys()) | set(met_metrics.keys()))
    for m in metric_names:
        pv = primary_metrics.get(m, np.nan)
        mv = met_metrics.get(m, np.nan)
        one[f"primary_{m}"] = pv
        one[f"met_{m}"] = mv
        if pd.notna(pv) and pd.notna(mv):
            one[f"diff_{m}"] = mv - pv
        else:
            one[f"diff_{m}"] = np.nan

    pair_rows.append(one)

    # pair plots
    pair_name = f"{int(patient_id)}_{primary_sample}_vs_{met_sample}"
    pair_dir = PAIR_PLOT_DIR / pair_name
    pair_dir.mkdir(parents=True, exist_ok=True)

    pair_density_plot(
        primary_raw, met_raw,
        pair_dir / "pair_density.png",
        title=f"{pair_name}: distance density"
    )
    pair_boxplot(
        primary_raw, met_raw,
        pair_dir / "pair_boxplot.png",
        title=f"{pair_name}: distance boxplot"
    )
    pair_within_r_plot(
        primary_metrics, met_metrics,
        pair_dir / "pair_within_r.png",
        title=f"{pair_name}: within-r"
    )

pair_df = pd.DataFrame(pair_rows)
if len(pair_df) == 0:
    raise RuntimeError("没有成功生成任何样本对结果，请检查第一段输出。")

# save pair_table
pair_table_path = OUT_DIR / "pair_table.tsv"
pair_df.to_csv(pair_table_path, sep="\t", index=False)

# =========================
# 5) COHORT-LEVEL ANALYSIS (All pairs)
# =========================
diff_cols = [c for c in pair_df.columns if c.startswith("diff_")]
cohort_rows = []

# 5.1 stats
for c in diff_cols:
    x = pair_df[c].to_numpy(dtype=float)
    x = x[~np.isnan(x)]
    if len(x) < 2:
        continue
    cohort_rows.append({
        "section": "stats",
        "metric": c.replace("diff_", ""),
        "n_pairs": int(len(x)),
        "diff_mean": float(np.mean(x)),
        "diff_median": float(np.median(x)),
        "diff_p25": float(np.quantile(x, 0.25)),
        "diff_p75": float(np.quantile(x, 0.75)),
        "wilcoxon_p": wilcoxon_safe(x),
        "x": np.nan,
        "y": np.nan,
        "n": np.nan,
        "spearman_rho": np.nan,
        "spearman_p": np.nan,
    })

# 5.2 correlations
qty_candidates = [c for c in diff_cols if c in ["diff_tumor_n", "diff_m1_n", "diff_m2_n", "diff_m2_frac_of_mac"]]
spatial_candidates = [c for c in diff_cols if any(k in c for k in ["dist_", "within_", "enrich_", "delta_median", "tumor_neighbor_"])]

for xc in qty_candidates:
    for yc in spatial_candidates:
        a = pair_df[xc].to_numpy(dtype=float)
        b = pair_df[yc].to_numpy(dtype=float)
        rho, p, n = spearman_safe(a, b)
        if np.isnan(rho) or n < 3:
            continue
        cohort_rows.append({
            "section": "correlation",
            "metric": np.nan,
            "n_pairs": np.nan,
            "diff_mean": np.nan,
            "diff_median": np.nan,
            "diff_p25": np.nan,
            "diff_p75": np.nan,
            "wilcoxon_p": np.nan,
            "x": xc.replace("diff_", ""),
            "y": yc.replace("diff_", ""),
            "n": n,
            "spearman_rho": rho,
            "spearman_p": p,
        })

cohort_df = pd.DataFrame(cohort_rows)
cohort_analysis_path = OUT_DIR / "cohort_analysis.tsv"
cohort_df.to_csv(cohort_analysis_path, sep="\t", index=False)

# =========================
# 6) STRATIFIED ANALYSIS
# =========================
pair = pair_df.copy()

# add grouping labels
pair["met_group"] = pair["met_sample"].apply(met_group)
pair["site_group"] = pair["met_sample"].apply(ovary_group)

# =========================
# NEW: sort rows by metastasis group, then by met_sample numeric order
# target order example:
#   5R, 6R, 7R, 8R, 13R, 16R, 21R
#   then L-only
#   then QT-only / others
# =========================
met_order_map = {
    "R-only": 0,
    "L-only": 1,
    "QT-only": 2,
    "Unknown": 3
}

pair["_met_order"] = pair["met_group"].map(met_order_map).fillna(99)
pair["_met_num"] = pair["met_sample"].apply(extract_sample_num)

pair = pair.sort_values(
    by=["_met_order", "_met_num", "met_sample", "patient_id"],
    ascending=[True, True, True, True]
).reset_index(drop=True)

pair = pair.drop(columns=["_met_order", "_met_num"])

# save annotated pair table
pair_out_tsv = OUT_DIR / "pair_table_with_group.tsv"
pair_out_xlsx = OUT_DIR / "pair_table_with_group.xlsx"
save_tsv_and_xlsx(pair, pair_out_tsv, pair_out_xlsx)

# key metrics to plot
key_metrics = [
    "m2_frac_of_mac",
    "within_20_M2",
    "within_30_M2",
    "enrich_20_Tumor_M1",
    "enrich_30_Tumor_M1",
    "tumor_neighbor_M1_ratio",
    "tumor_neighbor_M2_ratio",
]
key_metrics = [m for m in key_metrics if f"diff_{m}" in pair.columns]

stats_rows = []
corr_rows = []

def run_group(group_name: str, sub: pd.DataFrame):
    # stats
    for c in diff_cols:
        x = sub[c].to_numpy(dtype=float)
        x = x[~np.isnan(x)]
        if len(x) < 2:
            continue
        stats_rows.append({
            "group": group_name,
            "metric": c.replace("diff_", ""),
            "n_pairs": int(len(x)),
            "diff_mean": float(np.mean(x)),
            "diff_median": float(np.median(x)),
            "diff_p25": float(np.quantile(x, 0.25)),
            "diff_p75": float(np.quantile(x, 0.75)),
            "wilcoxon_p": wilcoxon_safe(x)
        })

    # correlations
    qty_candidates_local = [c for c in diff_cols if c in ["diff_tumor_n", "diff_m1_n", "diff_m2_n", "diff_m2_frac_of_mac"]]
    spatial_candidates_local = [c for c in diff_cols if any(k in c for k in ["dist_", "within_", "enrich_", "delta_median", "tumor_neighbor_"])]

    for xc in qty_candidates_local:
        for yc in spatial_candidates_local:
            a = sub[xc].to_numpy(dtype=float)
            b = sub[yc].to_numpy(dtype=float)
            rho, p, n = spearman_safe(a, b)
            if np.isnan(rho) or n < 3:
                continue
            corr_rows.append({
                "group": group_name,
                "x": xc.replace("diff_", ""),
                "y": yc.replace("diff_", ""),
                "n": n,
                "spearman_rho": rho,
                "spearman_p": p
            })

    # plots
    gdir = STRAT_PLOT_DIR / group_name.replace("/", "_")
    gdir.mkdir(parents=True, exist_ok=True)

    for m in key_metrics:
        paired_line_plot(
            sub, m,
            gdir / f"paired_{m}.png",
            title=f"{group_name}：{metric_cn(m)}配对变化（原发->转移）"
        )
        diff_hist(
            sub, m,
            gdir / f"diff_hist_{m}.png",
            title=f"{group_name}: diff distribution {m} (Met-Primary)"
        )

groups = {
    "All_pairs": pair,
    "Ovary(L/R)": pair[pair["site_group"] == "Ovary(L/R)"].copy(),
    "QT-only": pair[pair["site_group"] == "QT-only"].copy(),
    "L-only": pair[pair["met_group"] == "L-only"].copy(),
    "R-only": pair[pair["met_group"] == "R-only"].copy(),
}

print("\n================= 分层分析：组内配对数 =================")
for gname, sub in groups.items():
    print(f"- {gname:10s} : n_pairs = {len(sub)}")

for gname, sub in groups.items():
    if len(sub) < 2:
        print(f"[跳过] {gname} 配对数 < 2，无法做Wilcoxon")
        continue
    run_group(gname, sub)

stats_df = pd.DataFrame(stats_rows)
corr_df = pd.DataFrame(corr_rows)

group_order_map = {
    "R-only": 0,
    "L-only": 1,
    "QT-only": 2,
    "Ovary(L/R)": 3,
    "All_pairs": 4
}

if len(stats_df) > 0:
    stats_df["_group_order"] = stats_df["group"].map(group_order_map).fillna(99)
    stats_df = stats_df.sort_values(
        ["_group_order", "wilcoxon_p", "metric"]
    ).drop(columns=["_group_order"]).reset_index(drop=True)

if len(corr_df) > 0:
    corr_df["_group_order"] = corr_df["group"].map(group_order_map).fillna(99)
    corr_df = corr_df.sort_values(
        ["_group_order", "spearman_p"]
    ).drop(columns=["_group_order"]).reset_index(drop=True)

stats_tsv = OUT_DIR / "stratified_stats.tsv"
stats_xlsx = OUT_DIR / "stratified_stats.xlsx"
corr_tsv = OUT_DIR / "stratified_correlations.tsv"
corr_xlsx = OUT_DIR / "stratified_correlations.xlsx"

save_tsv_and_xlsx(stats_df, stats_tsv, stats_xlsx)
save_tsv_and_xlsx(corr_df, corr_tsv, corr_xlsx)

# =========================
# 7) TERMINAL HIGHLIGHTS
# =========================
print("\n================= 每组最显著指标 Top5（按p） =================")
if len(stats_df) > 0:
    for gname in [k for k in groups.keys() if k in stats_df["group"].unique()]:
        sub = stats_df[stats_df["group"] == gname].copy().sort_values("wilcoxon_p")
        if len(sub) == 0:
            continue
        print(f"\n--- {gname} ---")
        for _, r in sub.head(5).iterrows():
            direction = "Met>Primary" if r["diff_median"] > 0 else ("Met<Primary" if r["diff_median"] < 0 else "≈0")
            print(f"{r['metric']:22s} dir:{direction:10s} diff_med={fmt(r['diff_median'])} p={fmt(r['wilcoxon_p'])} n={int(r['n_pairs'])}")
else:
    print("(无可报告统计结果)")

print("\n================= 每组相关 Top5（按p） =================")
if len(corr_df) > 0:
    for gname in corr_df["group"].unique():
        sub = corr_df[corr_df["group"] == gname].copy().sort_values("spearman_p")
        print(f"\n--- {gname} ---")
        for _, r in sub.head(5).iterrows():
            print(f"{r['x']:18s} vs {r['y']:22s} rho={fmt(r['spearman_rho'])} p={fmt(r['spearman_p'])} n={int(r['n'])}")
else:
    print("(相关分析：配对数不足或数据缺失，未形成可报告结果)")

print("\n✅ 输出完成：")
print(f"- pair_table：{pair_table_path}")
print(f"- cohort_analysis：{cohort_analysis_path}")
print(f"- pair_table_with_group TSV ：{pair_out_tsv}")
print(f"- pair_table_with_group Excel：{pair_out_xlsx}")
print(f"- stratified_stats TSV      ：{stats_tsv}")
print(f"- stratified_stats Excel    ：{stats_xlsx}")
print(f"- stratified_corr TSV       ：{corr_tsv}")
print(f"- stratified_corr Excel     ：{corr_xlsx}")
print(f"- pair plots                ：{PAIR_PLOT_DIR}")
print(f"- stratified plots          ：{STRAT_PLOT_DIR}")
