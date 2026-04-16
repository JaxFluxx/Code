from openpyxl import Workbook,load_workbook
from openpyxl.formula.translate import Translator

wb = load_workbook(r'C:\Users\何嘉\Desktop\test.xlsx')
#ws = wb.active
ws_Sheet3 = wb['Sheet3']

ws_Sheet3.append(["价格1", "价格2", "总和", "平均值"])
ws_Sheet3.append([22, 63])
ws_Sheet3.append([11, 88])
ws_Sheet3.append([15, 68])

ws_Sheet3["c2"] = "=SUM(A2,B2)"
ws_Sheet3["d2"] = "=AVERAGE(A2:B2)"

# C3、C4使用上面的C2的求和公式
for cell in ws_Sheet3["C3:C4"]:
    # ws["C3"] = Translator(formula="=SUM(A2,B2)", origin="C2").translate_formula("C3")
    
    # 使用 Translator 类来翻译公式。
    cell[0].value = Translator(formula="=SUM(A2,B2)", origin="C2").translate_formula(cell[0].coordinate)
    # cell[0] 就是当前单元格对象。例如，第一次循环时，cell[0] 是 C3 单元格对象，第二次循环时，cell[0] 是 C4 单元格对象。
wb.save(r'C:\Users\何嘉\Desktop\test.xlsx')