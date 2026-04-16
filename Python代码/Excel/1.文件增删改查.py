from openpyxl import Workbook, load_workbook

#wb = Workbook()  # 创建一个工作簿对象

#打开工作簿
#  ws 是 wb 中的一个表格对象
wb = load_workbook(r'C:\Users\何嘉\Desktop\test.xlsx')  # 打开一个工作簿对象
ws = wb.active  # 获取当前活动表格,返回一个worksheet对象
print(wb.sheetnames)  # 输出工作簿中的所有表格名称

# 创建、获取工作
ws2 = wb.create_sheet('Sheet2')  # 创建一个新的表格对象
ws3 = wb.create_sheet('Sheet3')  # 创建一个新的表格对象
# 根据map获取sheet对象来操作表格
ws_test = wb["Sheet2"]
print(ws_test.title)

# 移动、删除工作簿
wb.move_sheet(ws3,-1)   # 将Sheet3移动到最后一个位置
del wb["Sheet3"]




wb.save(r'C:\Users\何嘉\Desktop\test.xlsx')