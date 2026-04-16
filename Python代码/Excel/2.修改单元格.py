from openpyxl import load_workbook, Workbook

wb = load_workbook(r'C:\Users\何嘉\Desktop\test.xlsx')
#ws = wb.active
ws_Sheet2 = wb['Sheet2']
print(wb.sheetnames)

#第一种
ws_Sheet2['A1'] = "你好"

#第二种
cell = ws_Sheet2.cell(2,1,"hello")
print(cell.value)

#行、列
print(cell.coordinate)  # A2
print(cell.row)  # 2
print(cell.column)  # 1

# 赋值
x = 1
for row in range(1,11):
    for column in range(1,11):
        ws_Sheet2.cell(row,column,x)
        print(ws_Sheet2.cell(row,column).value)
        x += 1

## 插入行/列
ws_Sheet2.insert_rows(1,2)  # 在第1行插入两行
ws_Sheet2.insert_cols(1,2)  # 在第1列插入两列

ws_Sheet2.append()

wb.save(r'C:\Users\何嘉\Desktop\test.xlsx')